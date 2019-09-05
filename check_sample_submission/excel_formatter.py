from openpyxl.styles import PatternFill

ERROR = 1
WARNING = 2
RED_FILL = PatternFill(start_color='FFFF0000',end_color='FFFF0000',fill_type='solid')


class ExcelFormatter:
    def __init__(self, sample_list_sheet, start_row_idx, sample_idx, volume_idx):
        self.sample_idx = sample_idx
        self.volume_idx = volume_idx
        self.sample_list_sheet = sample_list_sheet
        self.start_row_idx = start_row_idx
        self._reset_background()

    def mark_volume(self, relative_row_idx, level):
        color = None
        if level == ERROR:
            color = RED_FILL
        row_idx = relative_row_idx + self.start_row_idx
        self.sample_list_sheet.cell(
            row=row_idx + 1, column=self.volume_idx + 1).fill = color

    def mark_sample(self, relative_row_idx, level):
        color = None
        if level == ERROR:
            color = RED_FILL
        row_idx = relative_row_idx + self.start_row_idx
        self.sample_list_sheet.cell(
            row=row_idx + 1, column=self.sample_idx + 1).fill = color

    def _reset_background(self):
        no_fill = PatternFill(fill_type=None)
        for row in self.sample_list_sheet.iter_rows(min_row=self.start_row_idx + 1):
            for cell in row:
                if cell.fill != no_fill:
                    cell.fill = no_fill
