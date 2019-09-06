import openpyxl as xl
from check_sample_submission.utils import lazyprop
from check_sample_submission.sheet_validator import SheetValidator
from check_sample_submission.excel_formatter import ExcelFormatter


CAPTION_ROW_OFFSET = 5


class ImportSampleSubmission:
    def __init__(self, file_path, sample_column_index, volume_column_index,
                 sample_column_name, export_to_file_path=None, logger=None):
        self.logger = logger
        self.export_to_file_path = \
            file_path if export_to_file_path is None else export_to_file_path
        self.file_path = file_path
        self.sample_column_index = sample_column_index
        self.volume_column_index = volume_column_index
        self.sample_column_name = sample_column_name
        self.workbook = xl.load_workbook(self.file_path)
        self.sample_list_sheet = self.workbook['Sample list']

    def check_faulty_contents(self):
        start_row_idx = self._determine_caption_row_index() + CAPTION_ROW_OFFSET
        formatter = ExcelFormatter(
            self.sample_list_sheet, start_row_idx, self.sample_column_index,
            self.volume_column_index)
        validator = SheetValidator(excel_formatter=formatter, logger=self.logger)
        error_messages = validator.list_faulty(self.contents)
        return error_messages

    def save_workbook(self):
        try:
            self.workbook.save(self.export_to_file_path)
        except PermissionError:
            return False
        return True

    def _cell_value(self, row_index, col_index):
        return self.sample_list_sheet.cell(row=row_index + 1, column=col_index + 1).value

    def _determine_caption_row_index(self):
        for i in range(30):
            if self._cell_value(i, self.sample_column_index) == self.sample_column_name:
                return i
        raise ImportError('The caption \'{}\' was not found in column #{}'.format(
            self.sample_column_name, self.sample_column_index + 1))

    def _check_caption_row_offset(self, caption_row):
        # First row after offset should be filled in with a sample name
        # Row before the first sample row should contain NaN
        has_first_sample = not self._isna(
            caption_row + CAPTION_ROW_OFFSET, self.sample_column_index)
        if not has_first_sample:
            raise ImportError('Fatal error, could not locate first sample in sheet'
                              ', row: {}, col: {}'.format(
                caption_row + CAPTION_ROW_OFFSET + 1, self.sample_column_index + 1))

        is_cell_above_first_empty = self._isna(
            caption_row + CAPTION_ROW_OFFSET - 1, self.sample_column_index)
        if not is_cell_above_first_empty:
            raise ImportError('Fatal error, could not locate first sample in sheet, '
                              'cell is not empty, row: {}, col: {}'.format(
                caption_row + CAPTION_ROW_OFFSET, self.sample_column_index + 1))

    def _check_volume_column_exists(self):
        row = self._determine_caption_row_index()
        if self._cell_value(row, self.volume_column_index) != 'VOLUME':
            raise ImportError('The caption \'VOLUME\' was not found at column #{}'.format(
                self.volume_column_index + 1))

    def _fetch_contents(self):
        contents = list()
        self._check_volume_column_exists()
        caption_row_index = self._determine_caption_row_index()
        self._check_caption_row_offset(caption_row_index)
        for row in self.sample_list_sheet.iter_rows(min_row=caption_row_index + CAPTION_ROW_OFFSET + 1):
            sample_name = str(row[self.sample_column_index].value)
            volume = str(row[self.volume_column_index].value)
            if sample_name != 'None':
                contents.append((sample_name, volume))
        return contents

    @lazyprop
    def _samples(self):
        return [row[0] for row in self.contents]

    @lazyprop
    def contents(self):
        return self._fetch_contents()

    def _isna(self, row_index, col_index):
        return str(self._cell_value(row_index, col_index)) == 'None'
